#!/usr/bin/env python3
"""Build requirement-to-case execution traceability workbook."""

import argparse
import json

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


COLUMNS = [
    ("需求ID", 16, "requirement_id"),
    ("需求摘要", 42, "requirement_text"),
    ("关联用例数", 12, "case_count"),
    ("关联用例", 60, "case_names"),
    ("PASS", 10, "pass"),
    ("FAIL", 10, "fail"),
    ("N/A", 10, "na"),
    ("N/T", 10, "nt"),
    ("未执行", 10, "not_yet"),
    ("追踪状态", 14, "status"),
]


def main():
    parser = argparse.ArgumentParser(description="生成需求追踪矩阵 Excel")
    parser.add_argument("analysis_json")
    parser.add_argument("-o", "--output", required=True)
    args = parser.parse_args()
    with open(args.analysis_json, encoding="utf-8") as source:
        data = json.load(source)

    wb = Workbook()
    ws = wb.active
    ws.title = "需求追踪矩阵"
    thin = Side(style="thin", color="FFB7B7B7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    status_colors = {
        "PASS": "FF92D050",
        "FAIL": "FFFF6666",
        "PARTIAL": "FFFFC000",
        "NOT YET": "FFD9E1F2",
    }
    for column, (title, width, _) in enumerate(COLUMNS, start=1):
        cell = ws.cell(1, column, title)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="FFC4BD97")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
        ws.column_dimensions[cell.column_letter].width = width

    for row, item in enumerate(data.get("traceability", []), start=2):
        for column, (_, _, key) in enumerate(COLUMNS, start=1):
            value = item.get(key, "")
            if key == "case_names":
                value = "\n".join(value)
            cell = ws.cell(row, column, value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
            if key == "status" and value in status_colors:
                cell.fill = PatternFill("solid", fgColor=status_colors[value])

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:J{max(ws.max_row, 1)}"
    wb.save(args.output)
    print(f"✓ 已生成需求追踪矩阵 {args.output}：{len(data.get('traceability', []))} 条需求")


if __name__ == "__main__":
    main()
