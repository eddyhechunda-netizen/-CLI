#!/usr/bin/env python3
"""
build_testcase_xlsx.py — 把结构化用例数据（JSON）渲染成一份测试用例 .xlsx。

输出文件高保真复刻"需求→测试用例"标准模板：
  · 首页：项目元信息 + 目录汇总表（用 INDIRECT/COUNTIF 公式自动聚合各 sheet 统计）
  · 每个测试分类一个 sheet：左上统计区（COUNT/COUNTIF 公式 + 配色图例）、
    表头行、字段说明行、数据区（功能模块列按模块纵向合并）

设计意图：让每次生成都走同一条确定性渲染路径，使用例内容（人写/Claude 写）与
排版（脚本负责）彻底解耦。Claude 只需产出干净的 JSON，剩下的公式、合并、样式、
列宽全部由本脚本保证与模板一致，避免每次重新发明排版。

用法：
    python build_testcase_xlsx.py cases.json -o 输出.xlsx

JSON 结构见同目录 references/cases_schema.md，最小示例：
{
  "meta": {"project_id": "TRON2-V5", "version": "r-2.1.17", "testers": "@Eddy@Logen"},
  "sheets": [
    {
      "name": "硬件性能",
      "title": "硬件性能",
      "cases": [
        {"module": "硬件性能-充电桩机械结构",
         "name": "靠墙安装-贴墙安装（0-5cm）",
         "precondition": "充电桩未固定，墙面平整",
         "steps": "1.将充电桩贴墙放置\n2.手动模拟机器人对接3次",
         "expected": "充电桩不倾倒，晃动≤0.1cm",
         "type": "功能测试", "status": "正常", "level": "高", "creator": ""}
      ]
    }
  ]
}
"""
import argparse
import json
import math
import sys
import unicodedata

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ── 模板规格常量（从标准模板逐格提取，改这里即可整体调样式）──────────────

THIN = Side(style="thin", color="FF000000")
BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# 配色（ARGB，与模板一致）
C_HEADER = "FFC4BD97"   # 表头行 暖灰
C_DESC = "FFEBF1DE"     # 字段说明行 浅绿
C_PASS = "FF008000"     # 绿
C_FAIL = "FFFF0000"     # 红
C_NA = "FF808080"       # 灰  N/A
C_NT = "FFFFC000"       # 橙  N/T
C_TOTAL = "FFC3DD40"    # 首页 Total 行 黄绿
C_LINK = "FF0563C1"     # 首页 sheet 名超链接蓝

# 字体
F_STAT = Font(name="Calibri", size=10, bold=True, color="FF000000")    # 统计区/表头加粗
F_STAT_N = Font(name="Calibri", size=10, bold=False, color="FF000000")  # 统计区取值
F_HOME = Font(name="Calibri", size=10, bold=True, color="FF000000")    # 首页加粗
F_HOME_N = Font(name="Calibri", size=10, bold=False, color="FF000000")
F_LINK = Font(name="Calibri", size=10, bold=False, color=C_LINK)
F_DESC = Font(name="Calibri", size=9, bold=False, color="FF000000")     # 字段说明行
F_DATA = Font(name="宋体", size=10, bold=False, color="FF000000")       # 数据区正文
F_TITLE = Font(name="宋体", size=10, bold=False, color="FF000000")      # 分类标题与目标在线模板一致

AL_C = Alignment(horizontal="center", vertical="center", wrap_text=True)
AL_L = Alignment(horizontal="left", vertical="center", wrap_text=True)
AL_LT = Alignment(horizontal="left", vertical="top", wrap_text=True)
AL_V = Alignment(vertical="center", wrap_text=True)  # 数据正文：水平默认、垂直居中、自动换行
AL_DATA = Alignment(horizontal="left", vertical="top", wrap_text=True)

# 数据区表头（第 9 行）与列宽（Excel 字符宽）。
# 两种布局：
#   "slim"（默认，12列）—— 团队实际交付标准，K=测试结果、L=备注
#   "full"（14列）—— 额外含 测试数据 / 测试图片视频 两列，留给需要附证据的场景
# 前 10 列两种布局完全一致；差异只在 K 列之后。
COLUMNS_BASE = [
    ("序号", 4, "center"),
    ("功能模块", 17, "center"),
    ("用例名称", 30, "left"),
    ("前置条件", 35, "default"),
    ("用例步骤", 38, "default"),
    ("预期结果", 52, "default"),
    ("用例类型", 15, "default"),
    ("用例状态", 13, "default"),
    ("用例等级", 16, "default"),
    ("创建人", 13, "default"),
]
# K 列起的尾部列：测试结果列标题可被各 sheet 覆盖（真实文件里有"测试结果"
# "Tron2测试结果"等多种写法），备注列标题同理。
COL_RESULT = ("测试结果", 14, "default")
COL_NOTE = ("备注", 22, "default")
COL_TESTDATA = ("测试数据", 23, "default")
COL_MEDIA = ("测试图片/视频", 24, "default")


def columns_for(layout):
    """按布局返回列定义列表。"""
    if layout == "full":
        return COLUMNS_BASE + [COL_RESULT, COL_TESTDATA, COL_MEDIA, COL_NOTE]
    return COLUMNS_BASE + [COL_RESULT, COL_NOTE]  # slim 默认

# 字段说明行（第 10 行），逐列；空串表示该列不写说明。文案对齐真实交付文件。
FIELD_HINTS = {
    3: "“用例名称”为必填项。",
    4: "“前置条件”请填写合法文本。",
    5: "“用例步骤”请填写合法文本。",
    6: "“预期结果”请填写合法文本。",
    7: "“用例类型”请填写：功能测试、性能测试、安全性测试、其他。",
    8: "“用例状态”请填写：正常、待更新、已废弃。",
    9: "“用例等级”请填写：高、中、低。",
    10: "人员类型字段请填写人员的昵称。",
}

# 统计区左侧结果图例（B3:B7）与对应填充
RESULT_LEGEND = [
    ("PASS", C_PASS),
    ("FAIL", C_FAIL),
    ("N/A", C_NA),
    ("N/T", C_NT),
    ("NOT YET", None),
]
# 统计区右侧元信息标签（D2:D6）
META_LABELS = ["模块名称", "测试人员", "测试日期", "测试版本", "测试设备编号"]

HEADER_ROW = 9   # 表头所在行
HINT_ROW = 10    # 字段说明行
DATA_START = 11  # 数据起始行


def _fill(color):
    return PatternFill("solid", fgColor=color) if color else PatternFill()


def _display_width(value):
    return sum(
        2 if unicodedata.east_asian_width(char) in {"W", "F", "A"} else 1
        for char in str(value or "")
    )


def _wrapped_line_count(value, column_width):
    lines = str(value or "").splitlines() or [""]
    usable_width = max(1, int(column_width) - 2)
    return sum(max(1, math.ceil(_display_width(line) / usable_width)) for line in lines)


def _data_row_height(values, columns):
    line_count = max(
        _wrapped_line_count(value, columns[index][1])
        for index, value in enumerate(values)
    )
    return min(180, max(30, line_count * 17 + 6))


def build_data_sheet(ws, sheet, layout="slim"):
    """渲染一个测试分类 sheet：统计区 + 表头 + 说明行 + 数据区。

    layout: "slim"（12列，默认，团队交付标准）或 "full"（14列，含测试数据/图片列）。
    sheet 可带 result_header / note_header 覆盖 K/L 列标题（真实文件里测试结果列
    有"测试结果""Tron2测试结果"等多种写法）。
    """
    title = sheet.get("title") or sheet["name"]
    if not title.endswith("测试"):
        title = f"{title}测试"
    cases = sheet.get("cases", [])
    columns = columns_for(layout)
    ncol = len(columns)
    result_col = len(COLUMNS_BASE) + 1   # K 列序号（测试结果），统计公式锚定此列

    # 列宽
    for idx, (_, width, _) in enumerate(columns, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    if layout == "slim":
        for column in ("G", "H", "I", "J"):
            ws.column_dimensions[column].hidden = True

    # ── 统计区 A1:E7 ──
    # 标题行（行1）：跨 A~E 合并并水平+垂直居中，宋体加粗放大（醒目标题）
    ws["A1"] = title
    ws.merge_cells("A1:E1")
    ws["A1"].font = F_TITLE
    ws["A1"].alignment = AL_C
    ws["A1"].border = BORDER_ALL

    # A2 测试总数 / C2 公式；A3:A7 合并"结果"
    rcol = get_column_letter(result_col)  # 测试结果列字母（slim=K，full 也是 K）
    ws["A2"] = "测试总数"
    ws["C2"] = "=COUNT(A:A)"   # 数据区序号列计数 = 用例总数
    ws["D2"] = META_LABELS[0]
    ws.merge_cells("A2:B2")
    ws.merge_cells("A3:A7")
    ws["A3"] = "结果"
    for i, (label, color) in enumerate(RESULT_LEGEND):
        r = 3 + i
        ws.cell(r, 2, label).fill = _fill(color)
        # C 列统计公式：统计"测试结果"列中等于该图例值的单元格数
        ws.cell(r, 3, f"=COUNTIF({rcol}:{rcol},B{r})")
    for i, label in enumerate(META_LABELS[1:], start=3):
        ws.cell(i, 4, label)  # D3..D6

    # 统计区样式：A1:E7 全框线、加粗居中
    for r in range(2, 8):
        for c in range(1, 6):
            cell = ws.cell(r, c)
            cell.border = BORDER_ALL
            cell.alignment = AL_C
            # C2/C3.. 取值列不加粗，D 列标签加粗左对齐
            if c == 3:
                cell.font = F_STAT_N
                cell.alignment = AL_L
            elif c == 4:
                cell.font = F_STAT
                cell.alignment = AL_V
            else:
                cell.font = F_STAT
    ws["E4"].number_format = "yyyy/m/d"  # 测试日期格
    # ── 表头行（第 9 行）── K/L 列标题可被 sheet 覆盖
    header_names = [c[0] for c in columns]
    header_names[result_col - 1] = sheet.get("result_header") or COL_RESULT[0]
    header_names[-1] = sheet.get("note_header") or COL_NOTE[0]
    for idx, name in enumerate(header_names, start=1):
        align = columns[idx - 1][2]
        cell = ws.cell(HEADER_ROW, idx, name)
        cell.font = F_STAT
        cell.fill = _fill(C_HEADER)
        cell.border = BORDER_ALL
        cell.alignment = AL_L if align == "left" else AL_C

    # ── 字段说明行（第 10 行）──
    for c in range(1, ncol + 1):
        cell = ws.cell(HINT_ROW, c)
        cell.fill = _fill(C_DESC)
        cell.border = BORDER_ALL
        cell.font = F_DESC
        cell.alignment = AL_LT
        if c in FIELD_HINTS:
            cell.value = FIELD_HINTS[c]

    # ── 数据区（第 11 行起）──
    # 模块合并：记录每段连续相同 module 的行区间
    merge_runs = []  # (start_row, end_row)
    run_start = None
    prev_mod = object()

    for i, case in enumerate(cases):
        r = DATA_START + i
        mod = case.get("module", "") or ""
        row_vals = [
            i + 1,                                  # A 序号
            mod,                                    # B 功能模块
            case.get("name", ""),                   # C 用例名称
            case.get("precondition", ""),           # D 前置条件
            case.get("steps", ""),                  # E 用例步骤
            case.get("expected", ""),               # F 预期结果
            case.get("type", "功能测试"),           # G 用例类型
            case.get("status", "正常"),             # H 用例状态
            case.get("level", "中"),                # I 用例等级
            case.get("creator", "") or "",          # J 创建人
            case.get("result", "") or "",           # K 测试结果
        ]
        if layout == "full":
            row_vals += [
                case.get("test_data", "") or "",    # L 测试数据
                case.get("media", "") or "",        # M 测试图片/视频
                case.get("note", "") or "",         # N 备注
            ]
        else:
            row_vals.append(case.get("note", "") or "")  # L 备注
        for c, val in enumerate(row_vals, start=1):
            cell = ws.cell(r, c, val)
            cell.font = F_DATA
            cell.border = BORDER_ALL
            # 用例文本可能以 = + - @ 开头（如预期结果"=100%"），openpyxl 会
            # 误当公式写入，导致 Excel 打开报 #NAME? 或触发 CSV 注入。除序号(int)
            # 外的字符串值显式声明为文本，让内容原样呈现。
            if c != 1 and isinstance(val, str) and val[:1] in ("=", "+", "-", "@"):
                cell.data_type = "s"
            # 序号、功能模块居中；其余左/默认 + 垂直居中自动换行
            cell.alignment = AL_C if c in (1, 2) else AL_DATA
        # Excel/飞书不会可靠地为换行文本自动调整行高，显式按列宽估算，
        # 避免长前置条件、步骤和预期结果溢出到相邻行。
        ws.row_dimensions[r].height = _data_row_height(row_vals, columns)
        # 模块合并段落跟踪
        if mod != prev_mod:
            if run_start is not None and r - 1 > run_start:
                merge_runs.append((run_start, r - 1))
            run_start = r
            prev_mod = mod
    # 收尾最后一段
    if cases:
        last_row = DATA_START + len(cases) - 1
        if run_start is not None and last_row > run_start:
            merge_runs.append((run_start, last_row))

    # 执行功能模块列(B)纵向合并；合并后值居中靠上读起来更自然，但模板用居中
    for s, e in merge_runs:
        ws.merge_cells(start_row=s, start_column=2, end_row=e, end_column=2)

    ws.row_dimensions[HEADER_ROW].height = 30
    ws.row_dimensions[HINT_ROW].height = 48

    # 测试结果列使用固定枚举；导入飞书在线表格后可直接下拉选择。
    validation_end = max(DATA_START + len(cases) - 1, DATA_START + 500)
    result_range = f"{rcol}{DATA_START}:{rcol}{validation_end}"
    result_validation = DataValidation(
        type="list",
        formula1='"PASS,FAIL,N/A,N/T,NOT YET"',
        allow_blank=True,
    )
    result_validation.error = "请选择 PASS、FAIL、N/A、N/T 或 NOT YET"
    result_validation.errorTitle = "测试结果格式错误"
    result_validation.prompt = "请选择测试结果"
    result_validation.promptTitle = "测试结果"
    ws.add_data_validation(result_validation)
    result_validation.add(result_range)

    result_colors = {
        "PASS": C_PASS,
        "FAIL": C_FAIL,
        "N/A": C_NA,
        "N/T": C_NT,
    }
    for label, color in result_colors.items():
        ws.conditional_formatting.add(
            result_range,
            FormulaRule(
                formula=[f'{rcol}{DATA_START}="{label}"'],
                fill=_fill(color),
                font=Font(color="FFFFFFFF", bold=True),
            ),
        )

    # 目标团队模板不冻结分类页，保持在线表格的默认滚动体验。


def build_home_sheet(ws, meta, sheet_names):
    """渲染首页：项目元信息条 + 目录汇总表（INDIRECT 聚合各 sheet 统计）。"""
    # 列宽
    home_widths = {"A": 27, "B": 13, "C": 14, "D": 20, "E": 12,
                   "F": 10, "G": 10, "H": 14, "I": 12, "J": 32}
    for col, w in home_widths.items():
        ws.column_dimensions[col].width = w

    # A1 横幅
    ws.merge_cells("A1:J1")
    ws["A1"] = f"{meta.get('project_id', '')} 测试用例"
    ws["A1"].font = F_TITLE
    ws["A1"].alignment = AL_C

    # 第 2 行：项目元信息（标签/取值交替，部分合并）
    ws.merge_cells("A2:A3"); ws["A2"] = "项目编号"
    ws.merge_cells("B2:B3"); ws["B2"] = "项目编号"
    ws.merge_cells("C2:D3"); ws["C2"] = meta.get("project_id", "")
    ws.merge_cells("E2:E3"); ws["E2"] = "版本信息"
    ws.merge_cells("F2:G3"); ws["F2"] = meta.get("version", "")
    ws.merge_cells("H2:H3"); ws["H2"] = "测试人员"
    ws.merge_cells("I2:I3"); ws["I2"] = meta.get("testers", "")
    ws.merge_cells("J2:J3"); ws["J2"] = "测试时间"
    # A2 其实模板首格留作装饰，这里写"项目信息"更清楚
    ws["A2"] = "项目信息"

    # 第 4 行：目录表头
    ws.merge_cells("A4:A5"); ws["A4"] = "目录"
    ws.merge_cells("B4:B5"); ws["B4"] = "测试用例"
    ws.merge_cells("C4:C5"); ws["C4"] = "已测试"
    ws.merge_cells("D4:D5"); ws["D4"] = "用例完成率"
    ws.merge_cells("E4:E5"); ws["E4"] = "备注"
    ws.merge_cells("F4:J4"); ws["F4"] = "Maturity（成熟度）"
    # 第 5 行：成熟度图例 + 版本迭代记录
    legend = [("F5", "PASS\n测试通过", C_PASS), ("G5", "FAIL\n测试不通过", C_FAIL),
              ("H5", "NT\n暂无测试条件，后续补充", C_NT), ("I5", "NA\n无此功能", C_NA),
              ("J5", "版本迭代记录", None)]
    for coord, text, color in legend:
        ws[coord] = text
        if color:
            ws[coord].fill = _fill(color)

    # 第 6 行起：每个 sheet 一行，用 INDIRECT 引用其统计区
    # 注意：sheet 名可能含连字符/空格/括号（如"Pad-连接与设置"），Excel 公式引用
    # 这类 sheet 名必须用单引号包裹，否则 INDIRECT 解析失败返回 #REF!。
    # 故统一拼成 INDIRECT("'"&A6&"'!C2") 的形式。
    start = 6
    n = len(sheet_names)
    ws.merge_cells(f"J{start}:J{start + n - 1}")  # 版本迭代记录占一列纵向
    q = "\"'\"&"      # 前置单引号片段
    qe = "&\"'"       # 后置单引号片段（接 !Cx）
    for i, sn in enumerate(sheet_names):
        r = start + i
        # 飞书导入会丢弃 XLSX relationship 内部链接，公式链接可随单元格保留。
        escaped_sn = sn.replace("'", "''")
        ws.cell(
            r,
            1,
            f'=HYPERLINK("#\'{escaped_sn}\'!A1","{sn.replace(chr(34), chr(34) * 2)}")',
        ).font = F_LINK
        ws.cell(r, 2, f'=INDIRECT({q}A{r}{qe}!C2")')   # 测试用例总数
        ws.cell(r, 3, f"=F{r}+G{r}+H{r}+I{r}")          # 已测试 = PASS+FAIL+NT+NA（有结论的数）
        ws.cell(r, 4, f"=IFERROR(C{r}/B{r},0)")          # 完成率 = 已测/总数
        ws.cell(r, 4).number_format = "0.00%"
        ws.cell(r, 6, f'=INDIRECT({q}A{r}{qe}!C3")')   # PASS
        ws.cell(r, 7, f'=INDIRECT({q}A{r}{qe}!C4")')   # FAIL
        ws.cell(r, 8, f'=INDIRECT({q}A{r}{qe}!C6")')   # N/T
        ws.cell(r, 9, f'=INDIRECT({q}A{r}{qe}!C5")')   # N/A

    # Total 行
    tr = start + n
    ws.cell(tr, 1, "Total").fill = _fill(C_TOTAL)
    ws.cell(tr, 2, f"=SUM(B{start}:B{tr-1})")
    ws.cell(tr, 3, f"=SUM(C{start}:C{tr-1})")
    ws.cell(tr, 4, f"=IFERROR(C{tr}/B{tr},0)")
    ws.cell(tr, 4).number_format = "0.00%"
    ws.cell(tr, 6, f"=SUM(F{start}:F{tr-1})")
    ws.cell(tr, 7, f"=SUM(G{start}:G{tr-1})")
    ws.cell(tr, 8, f"=SUM(H{start}:H{tr-1})")
    ws.cell(tr, 9, f"=SUM(I{start}:I{tr-1})")

    # 统一套样式（边框+居中+加粗），覆盖到目录表全域 A1:J{tr}
    for r in range(1, tr + 1):
        for c in range(1, 11):
            cell = ws.cell(r, c)
            cell.border = BORDER_ALL
            if cell.alignment.horizontal is None:
                cell.alignment = AL_C
            # 仅给没显式设过字体的格补默认加粗
            if cell.font is None or (cell.font.color and cell.font.color.rgb == C_LINK):
                continue
            if not cell.font.bold and r != 1:
                cell.font = F_HOME

    # 行高
    home_heights = {1: 26, 2: 20, 3: 47, 4: 30, 5: 56}
    for r, h in home_heights.items():
        ws.row_dimensions[r].height = h
    for r in range(start, tr + 1):
        ws.row_dimensions[r].height = 40


def build_overview_sheet(ws, overview):
    """渲染「概述」sheet：核心功能扁平速览表。

    overview = {"title": "...", "items": [
        {"group": "双臂形态", "item": "开关机功能",
         "desc": "短按0.5s+长按3s开机...", "risk": "P0", "result": "PASS", "note": ""}, ...]}
    group 连续相同自动纵向合并；risk 用 P0/P1/P2。这是详细 sheet 的精简对照视图，
    方便管理者快速扫读"测了哪些核心项、结论如何"。
    """
    title = overview.get("title", "核心功能概述")
    items = overview.get("items", [])
    widths = {"A": 13, "B": 13, "C": 27, "D": 73, "E": 13, "F": 13, "G": 27}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    ws.merge_cells("A1:G1")
    ws["A1"] = title
    ws["A1"].font = F_TITLE
    ws["A1"].alignment = AL_C
    ws["A1"].border = BORDER_ALL
    ws.row_dimensions[1].height = 26

    headers = ["序号", "测试项", "子项", "测试描述", "风险等级", "结论", "备注"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(2, c, h)
        cell.font = F_STAT
        cell.fill = _fill(C_HEADER)
        cell.border = BORDER_ALL
        cell.alignment = AL_C

    result_fill = {"PASS": C_PASS, "FAIL": C_FAIL, "N/A": C_NA, "NA": C_NA, "N/T": C_NT}
    merge_runs, run_start, prev = [], None, object()
    for i, it in enumerate(items):
        r = 3 + i
        grp = it.get("group", "") or ""
        vals = [i + 1, grp, it.get("item", ""), it.get("desc", ""),
                it.get("risk", ""), it.get("result", ""), it.get("note", "")]
        for c, v in enumerate(vals, start=1):
            cell = ws.cell(r, c, v)
            cell.font = F_DATA
            cell.border = BORDER_ALL
            cell.alignment = AL_C if c in (1, 2, 5, 6) else AL_V
            if c != 1 and isinstance(v, str) and v[:1] in ("=", "+", "-", "@"):
                cell.data_type = "s"
            if c == 6 and v in result_fill:  # 结论列着色
                cell.fill = _fill(result_fill[v])
        if grp != prev:
            if run_start is not None and r - 1 > run_start:
                merge_runs.append((run_start, r - 1))
            run_start = r
            prev = grp
    if items:
        last = 3 + len(items) - 1
        if run_start is not None and last > run_start:
            merge_runs.append((run_start, last))
    for s, e in merge_runs:
        ws.merge_cells(start_row=s, start_column=2, end_row=e, end_column=2)
    ws.freeze_panes = "A3"


def build_trace_data_sheet(ws, sheets):
    """保存用例到需求的机器可读映射，供执行结果分析使用。"""
    headers = [
        "sheet_name", "case_number", "case_name", "module",
        "requirement_ids", "requirement_text",
    ]
    ws.append(headers)
    for sheet in sheets:
        for index, case in enumerate(sheet.get("cases", []), start=1):
            req_ids = case.get("requirement_ids", [])
            if isinstance(req_ids, str):
                req_ids = [req_ids]
            ws.append([
                sheet["name"],
                index,
                case.get("name", ""),
                case.get("module", ""),
                ",".join(str(item) for item in req_ids if str(item).strip()),
                case.get("requirement_text", ""),
            ])
    ws.sheet_state = "hidden"


def build_workbook(data):
    wb = Workbook()
    # 默认 sheet 改为首页
    home = wb.active
    home.title = "首页"

    meta = data.get("meta", {})
    layout = meta.get("layout", "slim")  # slim(12列,默认) | full(14列)
    sheets = data.get("sheets", [])
    sheet_names = [s["name"] for s in sheets]

    # 可选「概述」sheet：核心功能扁平速览（测试项/描述/风险等级/结论）
    overview = data.get("overview") if meta.get("include_overview") else None
    if overview:
        build_overview_sheet(wb.create_sheet(title="概述"), overview)

    for sheet in sheets:
        ws = wb.create_sheet(title=sheet["name"])
        build_data_sheet(ws, sheet, layout=layout)

    build_trace_data_sheet(wb.create_sheet(title="_追踪数据"), sheets)
    build_home_sheet(home, meta, sheet_names)
    # 首页置顶（概述紧随其后）
    wb.move_sheet("首页", -wb.index(home))
    if overview:
        wb.move_sheet("概述", 1 - wb.index(wb["概述"]))
    wb.active = 0
    return wb


def main():
    ap = argparse.ArgumentParser(description="把用例 JSON 渲染成测试用例 xlsx")
    ap.add_argument("cases_json", help="结构化用例数据 JSON 路径")
    ap.add_argument("-o", "--output", required=True, help="输出 xlsx 路径")
    args = ap.parse_args()

    with open(args.cases_json, encoding="utf-8") as f:
        data = json.load(f)

    n_sheets = len(data.get("sheets", []))
    n_cases = sum(len(s.get("cases", [])) for s in data.get("sheets", []))
    if n_sheets == 0:
        print("错误：JSON 中没有 sheets。", file=sys.stderr)
        sys.exit(1)

    wb = build_workbook(data)
    wb.save(args.output)
    print(f"✓ 已生成 {args.output}：{n_sheets} 个分类 sheet，共 {n_cases} 条用例")


if __name__ == "__main__":
    main()
