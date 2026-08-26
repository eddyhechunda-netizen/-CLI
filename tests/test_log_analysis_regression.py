import json
import os
import subprocess
import tempfile
import unittest
from datetime import datetime
from pathlib import Path
from unittest.mock import Mock, patch

from openpyxl import load_workbook

from bot import lark_test_bot as bot


MOTOR9_TRIGGER = """
2026-08-03 11:44:19.714 W/ethercat(mroslaunch)(627/627): motor9 something happened, statusword 0x1208 code 0x0
2026-08-03 11:44:19.726 E/ethercat(mroslaunch)(627/627): Motor = 65535, 0xf10b, Too many loss.
2026-08-03 11:44:19.753 I/ethercat(mroslaunch)(627/627): [88] 'EtherCAT error rx frames of port0': 255
2026-08-03 11:44:19.754 I/ethercat(mroslaunch)(627/627): [90] 'EtherCAT lost link cnt of port1': 0
2026-08-03 11:44:37.094 I/ethercat(mroslaunch)(1058/1058): Motor 9 (slave 11) enabled successfully.
""".strip()

MOTOR5_RECOVERY_FAILURE = """
2026-08-03 11:32:36.706 W/ethercat(mroslaunch)(635/635): motor9 something happened, statusword 0x1208 code 0x0
2026-08-03 11:32:36.712 E/ethercat(mroslaunch)(635/635): Motor = 65535, 0xf10b, Too many loss.
2026-08-03 11:32:53.892 E/ethercat(mroslaunch)(1058/1058): Failed to enable motor 5 (slave 6) statusword 0x12a1 code 0x7121
2026-08-03 11:33:11.014 E/ethercat(mroslaunch)(1058/1058): Failed to enable motor 5 (slave 6) statusword 0x12a1 code 0x7121
""".strip()

MOTOR10_RECOVERY_FAILURE = """
2026-08-01 10:01:37.733 W/ethercat(mroslaunch)(640/640): motor9 something happened, statusword 0x1208 code 0x0
2026-08-01 10:01:37.740 E/ethercat(mroslaunch)(640/640): Motor = 65535, 0xf10b, Too many loss.
2026-08-01 10:01:55.221 E/ethercat(mroslaunch)(1058/1058): Failed to enable motor 10 (slave 12) statusword 0x12a1 code 0x7121
2026-08-01 10:02:12.400 E/ethercat(mroslaunch)(1058/1058): Failed to enable motor 10 (slave 12) statusword 0x12a1 code 0x7121
""".strip()

ALL_MOTORS_LINK_DROP = "\n".join(
    [
        *[
            f"2026-07-31 10:21:01.792 W/ethercat(mroslaunch)(635/635): "
            f"motor{motor} something happened, statusword 0x0 code 0xf"
            for motor in range(1, 11)
        ],
        "2026-07-31 10:21:01.802 E/ethercat(mroslaunch)(635/635): Motor = 65535, 0xf10b, Too many loss.",
        "2026-07-31 10:21:15.831 I/ethercat(mroslaunch)(635/635): [8a] 'EtherCAT error rx frames of port1': 255",
        "2026-07-31 10:21:15.831 I/ethercat(mroslaunch)(635/635): [90] 'EtherCAT lost link cnt of port1': 1",
        "2026-07-31 10:21:15.832 I/ethercat(mroslaunch)(635/635): [92] 'EtherCAT lost link cnt of port3': 1",
    ]
)


class EvidenceGateRegressionTests(unittest.TestCase):
    def test_motor9_trigger_and_recovery_pass(self):
        answer = (
            "电机9（motor9/slave11）出现 statusword 0x1208 code 0x0，"
            "随后直接触发 Too many loss；帧错误非零，lost link 为0，主站重启后恢复。"
        )
        bot.validate_log_analysis_answer(MOTOR9_TRIGGER, answer)

    def test_motor9_trigger_cannot_be_reclassified_as_manual_power(self):
        answer = (
            "电机9 statusword 0x1208 code 0x0，"
            "最终根因为遥控器手动掉电并在主站重启后恢复。"
        )
        with self.assertRaisesRegex(RuntimeError, "不得改判为遥控器"):
            bot.validate_log_analysis_answer(MOTOR9_TRIGGER, answer)

    def test_motor5_recovery_failure_pass(self):
        answer = (
            "初始触发为电机9 statusword 0x1208 code 0x0 后出现 Too many loss；"
            "后续未恢复原因为电机5 statusword 0x12a1 code 0x7121 堵转保护。"
        )
        bot.validate_log_analysis_answer(MOTOR5_RECOVERY_FAILURE, answer)

    def test_motor10_recovery_failure_pass(self):
        answer = (
            "初始触发为电机9 statusword 0x1208 code 0x0 后出现 Too many loss；"
            "后续未恢复原因为电机10 statusword 0x12a1 code 0x7121 堵转保护。"
        )
        bot.validate_log_analysis_answer(MOTOR10_RECOVERY_FAILURE, answer)

    def test_motor_range_and_nonzero_lost_link_pass(self):
        answer = (
            "motor1~10 全部出现 statusword 0x0 code 0xf，随后 Too many loss；"
            "[90]/[92] lost link 非零，属于硬件链路瞬断，不能判为遥控器掉电。"
        )
        bot.validate_log_analysis_answer(ALL_MOTORS_LINK_DROP, answer)

    def test_nonzero_lost_link_rejects_all_zero_claim(self):
        answer = (
            "motor1~10 全部出现 statusword 0x0 code 0xf；"
            "所有 lost link 计数均为0。"
        )
        with self.assertRaisesRegex(RuntimeError, "存在非零 lost link"):
            bot.validate_log_analysis_answer(ALL_MOTORS_LINK_DROP, answer)


class PreprocessingRegressionTests(unittest.TestCase):
    def test_startup_ecm_transient_returns_normal_short_answer(self):
        source = "\n".join(
            [
                bot.ECM_DEEP_ANALYSIS_MARKER,
                "2026-08-03 10:35:41.716 E/snowball: >>>>>|ecm err|",
                "2026-08-03 10:35:43.026 I/ethercat: Motor 10 (slave 12) enabled successfully.",
                "2026-08-03 10:35:43.033 I/snowball: ethercat ok! ecm ok",
            ]
        )

        answer = bot.render_normal_log_analysis(source)

        self.assertEqual(
            "**✅ 结论：这是一个正常日志，无 EtherCAT 通信异常或电机故障。**",
            answer,
        )
        self.assertNotIn("异常与错误分析", answer)
        self.assertNotIn("风险与建议", answer)
        self.assertNotIn("EtherCAT 主站异常根因", answer)

    def test_real_fault_does_not_use_normal_short_answer(self):
        source = "\n".join(
            [
                bot.ECM_DEEP_ANALYSIS_MARKER,
                "2026-08-03 11:44:19.714 E/snowball: >>>>>|ecm err|",
                MOTOR9_TRIGGER,
                "2026-08-03 11:44:37.170 I/snowball: ethercat ok! ecm ok",
            ]
        )

        self.assertIsNone(bot.render_normal_log_analysis(source))

    def test_deep_analysis_keeps_evidence_without_embedding_knowledge_base(self):
        log = "\n".join(
            [
                "2026-08-03 11:44:19.714 W/ethercat(mroslaunch)(1/1): motor9 something happened, statusword 0x1208 code 0x0",
                "2026-08-03 11:44:19.726 E/ethercat(mroslaunch)(1/1): Motor = 65535, 0xf10b, Too many loss.",
                "2026-08-03 11:44:19.727 I/snowball(mroslaunch)(2/2): >>>diag ecm level:2 code:-1 msg:Too many loss.",
                "2026-08-03 11:44:19.727 E/snowball(mroslaunch)(2/2): >>>>>ecm exp",
                "2026-08-03 11:44:19.727 E/snowball(mroslaunch)(2/2): >>>>>|ecm err|",
                "2026-08-03 11:44:31.184 I/ethercat(mroslaunch)(3/3): Found 12 slaves.",
                "2026-08-03 11:44:37.094 I/ethercat(mroslaunch)(3/3): Motor 9 (slave 11) enabled successfully.",
                "2026-08-03 11:44:37.170 I/snowball(mroslaunch)(4/4): >>>diag ecm level:0 code:0 msg:ethercat ok!",
            ]
        )
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "sample.log.active"
            path.write_text(log, encoding="utf-8")
            source = bot.prepare_log_analysis_source(path)

        self.assertIn("motor9 something happened", source)
        self.assertIn("Motor 9 (slave 11) enabled successfully", source)
        self.assertNotIn("# EtherCAT 主站异常诊断知识库", source)
        self.assertNotIn("EtherCAT 主站异常判断依据（知识库", source)

    def test_single_pass_extraction_matches_legacy_extractors(self):
        log = "\n".join(
            [
                "2026-08-08 11:00:00.000 I/other(mroslaunch)(1/1): node_name: ethercat(mroslaunch)",
                "2026-08-08 11:00:00.100 I/snowball(mroslaunch)(1/1): state:ST_IDLE",
                "2026-08-08 11:00:00.200 W/ethercat(mroslaunch)(1/1): motor2 something happened, statusword 0x0 code 0xf",
                "2026-08-08 11:00:00.300 I/ethercat(mroslaunch)(1/1): [90] 'EtherCAT lost link cnt of port1': 1",
                "2026-08-08 11:00:00.400 I/snowball(mroslaunch)(1/1): ethercat ok!",
            ]
        )
        heartbeat = Mock()
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "sample.log"
            path.write_text(log, encoding="utf-8")
            extracted = bot.extract_log_evidence(path, heartbeat=heartbeat)
            legacy_snowball = bot._extract_node_lines(path, "snowball")
            legacy_ethercat = bot._extract_node_lines(path, "ethercat")
            legacy_diagnostic = bot._extract_ethercat_diagnostic_lines(path)

        self.assertEqual(legacy_snowball, extracted["snowball"])
        self.assertEqual(legacy_ethercat, extracted["ethercat"])
        self.assertEqual(legacy_diagnostic, extracted["diagnostic"])
        self.assertGreaterEqual(heartbeat.call_count, 2)

    def test_extraction_preserves_slave_block_header_for_attribution(self):
        log = "\n".join(
            [
                "2026-08-08 11:12:11.428 I/ethercat(mroslaunch)(1/1): EcApplication.cpp(1090) clk(1) Slave1",
                "2026-08-08 11:12:11.429 I/ethercat(mroslaunch)(1/1): EcApplication.cpp(1119) clk(2) [90] 'EtherCAT lost link cnt of port1': 1",
                "2026-08-08 11:12:11.430 I/ethercat(mroslaunch)(1/1): EcApplication.cpp(1090) clk(3) Slave6",
                "2026-08-08 11:12:11.431 I/ethercat(mroslaunch)(1/1): EcApplication.cpp(1116) clk(4) [8f] 'EtherCAT lost link cnt of port0': 1",
            ]
        )
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "sample.log"
            path.write_text(log, encoding="utf-8")
            extracted = bot.extract_log_evidence(path)
            legacy = bot._extract_ethercat_diagnostic_lines(path)

        diagnostic = extracted["diagnostic"]
        self.assertTrue(any(line.endswith("Slave1") for line in diagnostic))
        self.assertTrue(any(line.endswith("Slave6") for line in diagnostic))
        self.assertEqual(legacy, diagnostic)


class UnsupportedFeishuLinkTests(unittest.TestCase):
    SYNC = "https://cwjgfm21di.feishu.cn/sync/VL8IdjstWseEkabT6n1cLVSAngd"
    DOCX = "https://cwjgfm21di.feishu.cn/docx/ABCdef123456"

    def test_bare_sync_link_detected(self):
        hint = bot.detect_unsupported_feishu_link(self.SYNC)
        self.assertIsNotNone(hint)
        self.assertIn("同步块", hint)

    def test_sync_link_with_generate_prefix_detected(self):
        hint = bot.detect_unsupported_feishu_link(f"生成测试用例 {self.SYNC}")
        self.assertIsNotNone(hint)
        self.assertIn("同步块", hint)

    def test_supported_docx_link_not_flagged(self):
        self.assertIsNone(bot.detect_unsupported_feishu_link(self.DOCX))
        self.assertIsNone(
            bot.detect_unsupported_feishu_link(f"生成测试用例 {self.DOCX}")
        )

    def test_plain_text_without_link_not_flagged(self):
        self.assertIsNone(bot.detect_unsupported_feishu_link("帮我生成测试用例"))

    def test_sheets_and_base_links_detected(self):
        self.assertIn(
            "电子表格",
            bot.detect_unsupported_feishu_link(
                "https://x.feishu.cn/sheets/XYZ789"
            ),
        )
        self.assertIn(
            "多维表格",
            bot.detect_unsupported_feishu_link(
                "https://x.feishu.cn/base/QWE456"
            ),
        )


class HumanoidNodeSupportTests(unittest.TestCase):
    def test_mission_node_detected_as_primary_for_humanoid(self):
        log = "\n".join(
            [
                "2026-08-08 11:00:00.100 I/mission_engine(mroslaunch)(1/1): state:ST_IDLE SN:HU_D04A001",
                "2026-08-08 11:00:00.200 E/mission_engine(mroslaunch)(1/1): >>>>>|ecm err|",
                "2026-08-08 11:00:00.300 I/mission_engine(mroslaunch)(1/1): ethercat ok!",
            ]
        )
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "humanoid.log.active"
            path.write_text(log, encoding="utf-8")
            extracted = bot.extract_log_evidence(path)
            source = bot.prepare_log_analysis_source(path)

        self.assertEqual(extracted["primary_node"], "mission_engine")
        self.assertEqual(len(extracted["snowball"][0]), 3)
        self.assertIn("分析节点：mission_engine", source)

    def test_snowball_node_still_detected_for_tron(self):
        log = "\n".join(
            [
                "2026-08-08 11:00:00.100 I/snowball(mroslaunch)(1/1): state:ST_IDLE SN:SF_TRON2A",
                "2026-08-08 11:00:00.200 I/snowball(mroslaunch)(1/1): ethercat ok!",
            ]
        )
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "tron.log.active"
            path.write_text(log, encoding="utf-8")
            extracted = bot.extract_log_evidence(path)
            source = bot.prepare_log_analysis_source(path)

        self.assertEqual(extracted["primary_node"], "snowball")
        self.assertIn("分析节点：snowball", source)

    def test_missing_primary_node_error_lists_both_nodes(self):
        log = "2026-08-08 11:00:00.100 I/ethercat(mroslaunch)(1/1): Found 12 slaves."
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "noprimary.log.active"
            path.write_text(log, encoding="utf-8")
            with self.assertRaises(RuntimeError) as ctx:
                bot.prepare_log_analysis_source(path)
        self.assertIn("snowball", str(ctx.exception))
        self.assertIn("mission_engine", str(ctx.exception))


class HumanoidWorkflowTests(unittest.TestCase):
    SAFETY_LOG = "\n".join(
        [
            "2026-08-08 11:00:00.100 I/mission_engine(mroslaunch)(1/1): state:ST_IDLE SN:HU_D04A001",
            "2026-08-08 11:00:01.000 E/monitor(mroslaunch)(2/2): EthercatMonitor add ethercatCommunicationExp motor 5 action HALF_STAND",
            "2026-08-08 11:00:01.100 E/ethercat(mroslaunch)(3/3): motor5 something happened, statusword 0x1208 code 0x0",
            "2026-08-08 11:00:01.200 E/ethercat(mroslaunch)(3/3): Motor = 65535, 0xf10b, Too many loss.",
            "2026-08-08 11:00:02.000 I/monitor(mroslaunch)(2/2): DiagnosticValue name:ability/walk level:OK",
            "2026-08-08 11:00:02.100 I/monitor(mroslaunch)(2/2): DiagnosticValue name:ability_running message:stand",
            "2026-08-08 11:00:03.000 I/monitor(mroslaunch)(2/2): PeripheralMonitor bat_vol:48.2 battery:80 current:5.0",
        ]
    )
    NORMAL_LOG = "\n".join(
        [
            "2026-08-08 11:00:00.100 I/mission_engine(mroslaunch)(1/1): state:ST_IDLE SN:HU_D04A001",
            "2026-08-08 11:00:01.000 I/monitor(mroslaunch)(2/2): EthercatMonitor ethercatResetNormal",
            "2026-08-08 11:00:02.000 I/monitor(mroslaunch)(2/2): DiagnosticValue name:ability/walk level:OK",
            "2026-08-08 11:00:02.100 I/monitor(mroslaunch)(2/2): DiagnosticValue name:imu level:OK",
            "2026-08-08 11:00:03.000 I/monitor(mroslaunch)(2/2): PeripheralMonitor bat_vol:48.2 battery:80 current:5.0",
        ]
    )

    def _prepare(self, log):
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "humanoid.log.active"
            path.write_text(log, encoding="utf-8")
            return bot.prepare_log_analysis_source(path)

    def test_safety_anomaly_detected_and_triggers_deep_analysis(self):
        source = self._prepare(self.SAFETY_LOG)
        self.assertIn(bot.HUMANOID_ANALYSIS_MARKER, source)
        self.assertIn("EthercatMonitor 安全异常事件", source)
        self.assertIn("HALF_STAND", source)
        # 安全异常必须触发主站下钻深度分析区块。
        self.assertIn(bot.ECM_DEEP_ANALYSIS_MARKER, source)

    def test_no_safety_anomaly_skips_deep_analysis_but_keeps_state(self):
        source = self._prepare(self.NORMAL_LOG)
        self.assertIn(bot.HUMANOID_ANALYSIS_MARKER, source)
        self.assertIn("无 ethercat 安全异常", source)
        self.assertNotIn(bot.ECM_DEEP_ANALYSIS_MARKER, source)
        # 状态分析证据（DiagnosticValue / PeripheralMonitor）仍必须提供。
        self.assertIn("DiagnosticValue", source)
        self.assertIn("PeripheralMonitor", source)

    def test_ew_focus_list_flags_concentrated_node(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "humanoid.log.active"
            path.write_text(self.SAFETY_LOG, encoding="utf-8")
            extracted = bot.extract_humanoid_evidence(path)
        counts = extracted["ew_node_counts"]
        self.assertGreaterEqual(counts.get("ethercat", {}).get("E", 0), 2)
        self.assertTrue(extracted["has_safety_anomaly"])

    def test_scattered_ew_lines_excluded_from_raw_dump(self):
        # 一个集中出现节点（busnode 4 条 E/W ≥ 阈值）+ 两个零星节点（各 1 条）。
        lines = ["2026-08-08 11:00:00.100 I/mission_engine(m)(1/1): SN:HU_D04A001"]
        for i in range(4):
            lines.append(
                f"2026-08-08 11:00:0{i}.000 E/busnode(m)(3/3): busnode fault detail {i}"
            )
        lines.append("2026-08-08 11:00:05.000 W/loner_a(m)(4/4): scattered warn AAA")
        lines.append("2026-08-08 11:00:06.000 W/loner_b(m)(5/5): scattered warn BBB")
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "humanoid.log.active"
            path.write_text("\n".join(lines), encoding="utf-8")
            extracted = bot.extract_humanoid_evidence(path)
        block = bot.build_humanoid_analysis_block(extracted)
        # 重点节点原文保留。
        self.assertIn("busnode fault detail", block)
        # 零星节点原文不进 dump，避免浪费 token。
        self.assertNotIn("scattered warn AAA", block)
        self.assertNotIn("scattered warn BBB", block)
        # 零星节点仍以计数形式体现。
        self.assertIn("loner_a", block)
        self.assertIn("loner_b", block)

    def test_diagnostic_value_change_point_compression(self):
        # imu 每周期重复 OK（应折叠为 1 条）；ability_running 发生 stand→walk 变化（应各留）。
        lines = ["2026-08-08 11:00:00.100 I/mission_engine(m)(1/1): SN:HU_D04A001"]
        for i in range(6):
            lines.append(
                f"2026-08-08 11:00:0{i}.000 I/monitor(m)(2/2): DiagnosticValue name:imu level:OK"
            )
        lines.append(
            "2026-08-08 11:00:00.500 I/monitor(m)(2/2): DiagnosticValue name:ability_running message:stand"
        )
        lines.append(
            "2026-08-08 11:00:05.500 I/monitor(m)(2/2): DiagnosticValue name:ability_running message:walk"
        )
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "humanoid.log.active"
            path.write_text("\n".join(lines), encoding="utf-8")
            extracted = bot.extract_humanoid_evidence(path)
        kept, name_count = bot._compress_diagnostic_lines(extracted["diagnostic"][0])
        imu_lines = [ln for ln in kept if "name:imu" in ln]
        run_lines = [ln for ln in kept if "ability_running" in ln]
        self.assertEqual(len(imu_lines), 1)  # 6 条重复折叠为 1
        self.assertEqual(len(run_lines), 2)  # stand / walk 两次变化都保留
        self.assertEqual(name_count, 2)

    def test_peripheral_downsample_flags_voltage_drop(self):
        lines = ["2026-08-08 11:00:00.100 I/mission_engine(m)(1/1): SN:HU_D04A001"]
        lines.append(
            "2026-08-08 11:00:00.000 I/monitor(m)(2/2): PeripheralMonitor bat_vol:48.0 battery:80 current:5.0"
        )
        # 骤降 >2V，应标记电源异常。
        lines.append(
            "2026-08-08 11:00:01.000 I/monitor(m)(2/2): PeripheralMonitor bat_vol:45.0 battery:78 current:5.0"
        )
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "humanoid.log.active"
            path.write_text("\n".join(lines), encoding="utf-8")
            extracted = bot.extract_humanoid_evidence(path)
        kept, anomalies = bot._compress_peripheral_lines(extracted["peripheral"][0])
        self.assertEqual(anomalies, 1)
        self.assertTrue(any("⚠电源异常" in ln for ln in kept))

    def test_tron_snowball_log_has_no_humanoid_block(self):
        log = "\n".join(
            [
                "2026-08-08 11:00:00.100 I/snowball(mroslaunch)(1/1): state:ST_IDLE SN:SF_TRON2A",
                "2026-08-08 11:00:01.000 I/monitor(mroslaunch)(2/2): DiagnosticValue name:imu level:OK",
            ]
        )
        source = self._prepare(log)
        self.assertNotIn(bot.HUMANOID_ANALYSIS_MARKER, source)


class CorrectionLoopRegressionTests(unittest.TestCase):
    class FakeProcess:
        def __init__(self, answer):
            self.answer = answer
            self.returncode = 0

        def communicate(self, timeout=None):
            return self.answer, ""

    def test_revision_reuses_session_and_sends_only_gate_error(self):
        calls = []
        prompts = []
        answers = iter(["incomplete answer", "corrected answer"])

        def fake_popen(args, **kwargs):
            calls.append(args)
            stdin = kwargs.get("stdin")
            if stdin is not None:
                stdin.seek(0)
                prompts.append(stdin.read())
            else:
                prompts.append("")
            return self.FakeProcess(next(answers))

        def fake_validate(_source, answer):
            if answer == "incomplete answer":
                raise RuntimeError("missing motor9")

        patches = (
            patch.object(bot, "start_ai_usage"),
            patch.object(bot, "finish_ai_usage"),
            patch.object(bot, "is_cancel_requested", return_value=False),
            patch.object(
                bot,
                "extract_log_evidence",
                return_value={
                    "snowball": (["snowball"], 1, False),
                    "ethercat": (["ethercat"], 1, False),
                    "diagnostic": [],
                },
            ),
            patch.object(bot, "prepare_log_analysis_source", return_value="evidence"),
            patch.object(bot, "build_prompt", return_value="FULL ORIGINAL PROMPT"),
            patch.object(bot, "set_job_progress"),
            patch.object(bot, "safe_update_job_card"),
            patch.object(bot, "validate_log_analysis_answer", side_effect=fake_validate),
            patch.object(bot.subprocess, "Popen", side_effect=fake_popen),
            patch.object(bot, "LOG_ANALYSIS_MAX_REVISIONS", 2),
        )
        with tempfile.TemporaryDirectory() as tmpdir, \
                patches[0], patches[1], patches[2], patches[3], patches[4], patches[5], \
                patches[6], patches[7], patches[8], patches[9], patches[10]:
            result = bot.run_copilot(
                {"job_id": "job-test", "action": "log_analysis", "source": "x"},
                tmpdir,
            )

        self.assertEqual("corrected answer", result)
        self.assertEqual(2, len(calls))
        session_id = calls[0][calls[0].index("--session-id") + 1]
        self.assertIn(f"--resume={session_id}", calls[1])
        # prompt 现在通过 stdin 传入（绕过命令行参数长度上限），不再使用 -p。
        self.assertNotIn("-p", calls[1])
        revision_prompt = prompts[1]
        self.assertIn("missing motor9", revision_prompt)
        self.assertNotIn("FULL ORIGINAL PROMPT", revision_prompt)
        self.assertNotIn("incomplete answer", revision_prompt)


class OperationsRegressionTests(unittest.TestCase):
    def test_progress_mapping_covers_live_stages(self):
        self.assertEqual(15, bot.job_stage_progress("读取日志", "running"))
        self.assertEqual(55, bot.job_stage_progress("分析日志", "running"))
        self.assertEqual(75, bot.job_stage_progress("校正结论", "running"))
        self.assertEqual(96, bot.job_stage_progress("更新边框", "running"))
        self.assertEqual(98, bot.job_stage_progress("更新交互", "running"))
        self.assertEqual(100, bot.job_stage_progress("分析日志", "done"))

    def test_python_stage_is_healthy_when_worker_heartbeat_is_alive(self):
        healthy, startup_grace = bot.running_job_health(
            heartbeat_age=20,
            process_running=False,
            worker_running=True,
            startup_age=120,
        )

        self.assertTrue(healthy)
        self.assertFalse(startup_grace)

    def test_two_calendar_day_retention_starts_at_yesterday_midnight(self):
        now = datetime(2026, 8, 13, 11, 30).timestamp()
        with patch.object(bot, "DATA_RETENTION_DAYS", 2):
            cutoff = datetime.fromtimestamp(bot.retention_cutoff_timestamp(now))

        self.assertEqual(datetime(2026, 8, 12, 0, 0), cutoff)

    def test_orphan_cache_cleanup_keeps_recent_and_registered_dirs(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            jobs_dir = Path(tmpdir)
            old_orphan = jobs_dir / "job_old"
            old_registered = jobs_dir / "job_registered"
            recent_orphan = jobs_dir / "_staging_recent"
            for path in (old_orphan, old_registered, recent_orphan):
                path.mkdir()
            cutoff = datetime(2026, 8, 12, 0, 0).timestamp()
            old_time = datetime(2026, 8, 11, 12, 0).timestamp()
            recent_time = datetime(2026, 8, 12, 12, 0).timestamp()
            os.utime(old_orphan, (old_time, old_time))
            os.utime(old_registered, (old_time, old_time))
            os.utime(recent_orphan, (recent_time, recent_time))

            with patch.object(bot, "JOBS_DIR", jobs_dir):
                removed = bot.cleanup_orphan_job_dirs(
                    cutoff,
                    {"job_registered"},
                )

            self.assertEqual(1, removed)
            self.assertFalse(old_orphan.exists())
            self.assertTrue(old_registered.exists())
            self.assertTrue(recent_orphan.exists())


class DeterministicCasePipelineTests(unittest.TestCase):
    def test_prefetch_preserves_complete_xml(self):
        source_xml = (
            "<title>需求</title><table><tr><td>触发时延 ≤ 50ms</td></tr></table>"
            '<img alt="碰撞保护示意图"/>'
        )
        payload = {"data": {"document": {"content": source_xml}}}
        with tempfile.TemporaryDirectory() as tmpdir, patch.object(
            bot, "run_json", return_value=payload
        ) as run_json_mock:
            path = bot.prefetch_requirement_source(
                "https://example.feishu.cn/wiki/requirement",
                Path(tmpdir),
            )
            self.assertEqual(source_xml, path.read_text(encoding="utf-8"))

        args = run_json_mock.call_args.args[0]
        self.assertEqual("xml", args[args.index("--doc-format") + 1])

    def test_cases_prompt_uses_local_source_without_upload_steps(self):
        prompt = bot.build_prompt(
            {
                "action": "cases",
                "source": "https://example.feishu.cn/wiki/requirement",
            },
            Path("/tmp/job"),
            "完整需求正文已保存到：/tmp/job/requirement_source.xml",
        )

        self.assertIn("/tmp/job/requirement_source.xml", prompt)
        self.assertIn("不要生成或上传最终产物", prompt)
        self.assertNotIn("sheets +workbook-import", prompt)
        self.assertNotIn("docs +create", prompt)

    def test_python_finalizes_excel_sheet_and_mindmap(self):
        example = bot.ROOT / "assets" / "example_cases.json"
        cases = json.loads(example.read_text(encoding="utf-8"))
        expected_count = sum(
            len(sheet.get("cases", []))
            for sheet in cases.get("sheets", [])
        )

        def fake_run_case_command(
            args,
            job_dir,
            _job_id,
            timeout=180,
            parse_json=False,
        ):
            if args[0] == "python3":
                result = subprocess.run(
                    args,
                    cwd=job_dir,
                    capture_output=True,
                    text=True,
                    timeout=timeout,
                )
                if result.returncode != 0:
                    self.fail(result.stderr or result.stdout)
                return result.stdout
            if "+workbook-import" in args:
                return {"data": {"url": "https://example.feishu.cn/sheets/sheet1"}}
            if "+create" in args:
                content = args[args.index("--content") + 1]
                doc_id = "quality1" if "质量检查" in content else "mindmap1"
                return {
                    "data": {
                        "document": {
                            "url": f"https://example.feishu.cn/docx/{doc_id}"
                        }
                    }
                }
            self.fail(f"unexpected lark command: {args}")

        with tempfile.TemporaryDirectory() as tmpdir:
            job_dir = Path(tmpdir)
            (job_dir / "cases.json").write_text(
                json.dumps(cases, ensure_ascii=False),
                encoding="utf-8",
            )
            (job_dir / "requirement_source.xml").write_text(
                "<title>TRON2 柔顺控制需求 V1.0</title>",
                encoding="utf-8",
            )
            (job_dir / "quality_review.json").write_text(
                json.dumps(
                    {
                        "meta": {"title": "质量检查"},
                        "requirements": [
                            {
                                "id": "REQ-001",
                                "module": "柔顺控制",
                                "text": "支持柔顺控制",
                                "testable": True,
                                "risk": "高",
                            }
                        ],
                    },
                    ensure_ascii=False,
                ),
                encoding="utf-8",
            )
            with patch.object(bot, "run_coverage_gate", return_value=("pass", "")), \
                    patch.object(
                        bot,
                        "_run_case_command",
                        side_effect=fake_run_case_command,
                    ), \
                    patch.object(bot, "set_job_progress"), \
                    patch.object(bot, "safe_update_job_card"):
                result = bot.finalize_case_artifacts(
                    {"job_id": "job-cases", "action": "cases"},
                    job_dir,
                    "结构化用例已生成。",
                )

            self.assertTrue((job_dir / "柔顺控制测试用例.xlsx").exists())
            self.assertTrue((job_dir / "柔顺控制测试点思维导图.xml").exists())
            self.assertTrue((job_dir / "柔顺控制需求质量检查报告.xml").exists())
            workbook = load_workbook(next(job_dir.glob("*.xlsx")))
            data_sheet = workbook[workbook.sheetnames[1]]
            heights = [
                data_sheet.row_dimensions[row].height
                for row in range(11, data_sheet.max_row + 1)
            ]
            self.assertTrue(all(height >= 30 for height in heights))
            self.assertTrue(any(height > 30 for height in heights))

        self.assertIn(f"测试用例：{expected_count} 条", result)
        self.assertIn("/sheets/sheet1", result)
        self.assertIn("/docx/quality1", result)
        self.assertIn("/docx/mindmap1", result)

    def test_case_artifact_buttons_are_distinct(self):
        result = """
- [柔顺控制测试用例](https://example.feishu.cn/sheets/sheet1)
- [柔顺控制需求质量检查报告](https://example.feishu.cn/docx/quality1)
- [柔顺控制测试点思维导图](https://example.feishu.cn/docx/mindmap1)
""".strip()
        artifacts = bot.parse_result_artifacts(result, "cases")

        self.assertEqual(
            ["打开测试用例", "打开质量检查报告", "打开测试点思维导图"],
            [item["button_label"] for item in artifacts],
        )


class DocumentQaRegressionTests(unittest.TestCase):
    def test_prompt_uses_job_instruction_as_question(self):
        prompt = bot.build_prompt(
            {
                "action": "doc_qa",
                "source": "https://example.feishu.cn/docx/document",
                "instruction": "这个需求的碰撞保护触发时延是多少？",
            },
            Path("/tmp/job"),
            "文档标题：测试需求\n文档正文：触发时延不超过50ms。",
        )

        self.assertIn("这个需求的碰撞保护触发时延是多少？", prompt)
        self.assertIn("触发时延不超过50ms", prompt)

    def test_citations_use_feishu_fragment_and_table_anchor(self):
        content = (
            '<h2 id="heading1">模式切换</h2>'
            '<table id="table1"><tr><td><p id="paragraph1">正序切换</p></td></tr></table>'
        )
        result = (
            "- [模式正序]"
            "(https://example.feishu.cn/wiki/doc1?blockId=paragraph1)"
        )

        remapped = bot.remap_doc_qa_citations(result, content)

        self.assertIn("https://example.feishu.cn/wiki/doc1#table1", remapped)
        self.assertNotIn("?blockId=", remapped)


if __name__ == "__main__":
    unittest.main()
