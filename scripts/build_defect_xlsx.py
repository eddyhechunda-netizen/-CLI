#!/usr/bin/env python3
"""Build a structured defect workbook from execution analysis JSON."""

import argparse
import json

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


COLUMNS = [
    ("缺陷ID", 12, "defect_id"),
    ("缺陷标题", 36, "title"),
    ("需求ID", 18, "requirement_ids"),
    ("功能模块", 20, "module"),
    ("来源用例", 30, "case_name"),
    ("前置条件", 28, "precondition"),
    ("复现步骤", 40, "steps"),
    ("预期结果", 36, "expected"),
    ("实际结果", 36, "actual"),
    ("严重程度", 12, "severity"),
    ("优先级", 10, "priority"),
    ("重现规律", 12, "reproducibility"),
    ("软件版本", 16, "software_version"),
    ("主站版本", 16, "station_version"),
    ("分电板版本", 16, "subboard_version"),
    ("驱动器版本", 16, "driver_version"),
    ("整机版本", 16, "machine_form"),
    ("来源Sheet", 18, "source_sheet"),
    ("来源行", 10, "source_row"),
]


def main():
    parser = argparse.ArgumentParser(description="生成结构化缺陷清单 Excel")
    parser.add_argument("analysis_json")
    parser.add_argument("-o", "--output", required=True)
    args = parser.parse_args()
    with open(args.analysis_json, encoding="utf-8") as source:
        data = json.load(source)

    wb = Workbook()
    ws = wb.active
    ws.title = "缺陷清单"
    thin = Side(style="thin", color="FFB7B7B7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for column, (title, width, _) in enumerate(COLUMNS, start=1):
        cell = ws.cell(1, column, title)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="FFC4BD97")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
        ws.column_dimensions[cell.column_letter].width = width

    for row, defect in enumerate(data.get("defects", []), start=2):
        for column, (_, _, key) in enumerate(COLUMNS, start=1):
            value = defect.get(key, "")
            if key == "requirement_ids":
                value = ",".join(value)
            cell = ws.cell(row, column, value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:S{max(ws.max_row, 1)}"
    wb.save(args.output)
    print(f"✓ 已生成缺陷清单 {args.output}：{len(data.get('defects', []))} 条")


if __name__ == "__main__":
    main()
