#!/usr/bin/env python3
"""Analyze a generated test-case workbook after execution."""

import argparse
import json
import sys
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook


RESULT_ALIASES = {
    "PASS": "PASS",
    "FAIL": "FAIL",
    "N/A": "N/A",
    "NA": "N/A",
    "N/T": "N/T",
    "NT": "N/T",
    "NOT YET": "NOT YET",
}


def normalize_result(value):
    raw = str(value or "").strip().upper()
    return RESULT_ALIASES.get(raw, raw or "NOT YET")


def split_ids(value):
    return [item.strip() for item in str(value or "").split(",") if item.strip()]


def load_trace_map(workbook):
    if "_追踪数据" not in workbook.sheetnames:
        return {}
    ws = workbook["_追踪数据"]
    headers = {str(cell.value): index for index, cell in enumerate(ws[1], start=1)}
    trace = {}
    for row in range(2, ws.max_row + 1):
        sheet_name = ws.cell(row, headers["sheet_name"]).value
        case_number = ws.cell(row, headers["case_number"]).value
        case_name = ws.cell(row, headers["case_name"]).value
        trace[(str(sheet_name), int(case_number or 0), str(case_name or ""))] = {
            "module": str(ws.cell(row, headers["module"]).value or ""),
            "requirement_ids": split_ids(ws.cell(row, headers["requirement_ids"]).value),
            "requirement_text": str(
                ws.cell(row, headers["requirement_text"]).value or ""
            ),
        }
    return trace


def find_column(ws, predicate):
    for column in range(1, ws.max_column + 1):
        value = str(ws.cell(9, column).value or "")
        if predicate(value):
            return column
    return None


def severity_for(level):
    return {"高": "严重", "中": "一般", "低": "提示"}.get(str(level), "一般")


def priority_for(level):
    return {"高": "高", "中": "中", "低": "低"}.get(str(level), "中")


def parse_workbook(path):
    wb = load_workbook(path, data_only=False)
    trace_map = load_trace_map(wb)
    home = wb["首页"] if "首页" in wb.sheetnames else None
    meta = {
        "project": str(home["C2"].value or "") if home else "",
        "version": str(home["F2"].value or "") if home else "",
        "testers": str(home["I2"].value or "") if home else "",
        "source_file": str(path),
    }

    cases = []
    for ws in wb.worksheets:
        if ws.title in {"首页", "概述", "_追踪数据"}:
            continue
        result_col = find_column(ws, lambda value: "测试结果" in value)
        note_col = find_column(
            ws,
            lambda value: value in {"备注", "记录", "测试问题记录"} or "备注" in value,
        )
        if not result_col:
            continue
        for row in range(11, ws.max_row + 1):
            case_number = ws.cell(row, 1).value
            case_name = ws.cell(row, 3).value
            if case_number in (None, "") or not str(case_name or "").strip():
                continue
            key = (ws.title, int(case_number), str(case_name))
            trace = trace_map.get(key, {})
            case = {
                "sheet": ws.title,
                "row": row,
                "case_number": int(case_number),
                "module": str(ws.cell(row, 2).value or trace.get("module", "")),
                "name": str(case_name),
                "precondition": str(ws.cell(row, 4).value or ""),
                "steps": str(ws.cell(row, 5).value or ""),
                "expected": str(ws.cell(row, 6).value or ""),
                "type": str(ws.cell(row, 7).value or ""),
                "level": str(ws.cell(row, 9).value or ""),
                "result": normalize_result(ws.cell(row, result_col).value),
                "note": str(ws.cell(row, note_col).value or "") if note_col else "",
                "requirement_ids": trace.get("requirement_ids", []),
                "requirement_text": trace.get("requirement_text", ""),
            }
            cases.append(case)

    if not cases:
        sys.exit("未找到可分析的测试用例，请确认文件由本 Skill 生成且表头未被删除。")

    counts = Counter(case["result"] for case in cases)
    concluded = sum(counts.get(value, 0) for value in ("PASS", "FAIL", "N/A", "N/T"))
    summary = {
        "total": len(cases),
        "tested": concluded,
        "completion_rate": round(concluded / len(cases), 4),
        "pass": counts.get("PASS", 0),
        "fail": counts.get("FAIL", 0),
        "na": counts.get("N/A", 0),
        "nt": counts.get("N/T", 0),
        "not_yet": counts.get("NOT YET", 0),
        "pass_rate": round(counts.get("PASS", 0) / concluded, 4) if concluded else 0,
    }

    defects = []
    for index, case in enumerate((item for item in cases if item["result"] == "FAIL"), start=1):
        defects.append({
            "defect_id": f"BUG-{index:03d}",
            "title": f"[{case['module']}] {case['name']}执行失败",
            "requirement_ids": case["requirement_ids"],
            "module": case["module"],
            "case_name": case["name"],
            "precondition": case["precondition"],
            "steps": case["steps"],
            "expected": case["expected"],
            "actual": case["note"] or "测试结果为FAIL，实际现象待补充。",
            "severity": severity_for(case["level"]),
            "priority": priority_for(case["level"]),
            "reproducibility": "",
            "software_version": meta["version"],
            "station_version": "",
            "subboard_version": "",
            "driver_version": "",
            "machine_form": "",
            "source_sheet": case["sheet"],
            "source_row": case["row"],
        })

    by_requirement = defaultdict(list)
    requirement_text = {}
    for case in cases:
        for requirement_id in case["requirement_ids"]:
            by_requirement[requirement_id].append(case)
            if case["requirement_text"]:
                requirement_text[requirement_id] = case["requirement_text"]

    traceability = []
    for requirement_id in sorted(by_requirement):
        linked = by_requirement[requirement_id]
        results = Counter(case["result"] for case in linked)
        if results.get("FAIL"):
            status = "FAIL"
        elif results.get("PASS") == len(linked):
            status = "PASS"
        elif all(case["result"] == "NOT YET" for case in linked):
            status = "NOT YET"
        else:
            status = "PARTIAL"
        traceability.append({
            "requirement_id": requirement_id,
            "requirement_text": requirement_text.get(requirement_id, ""),
            "case_count": len(linked),
            "case_names": [case["name"] for case in linked],
            "pass": results.get("PASS", 0),
            "fail": results.get("FAIL", 0),
            "na": results.get("N/A", 0),
            "nt": results.get("N/T", 0),
            "not_yet": results.get("NOT YET", 0),
            "status": status,
        })

    return {
        "meta": meta,
        "summary": summary,
        "cases": cases,
        "defects": defects,
        "traceability": traceability,
    }


def main():
    parser = argparse.ArgumentParser(description="分析已执行的测试用例 Excel")
    parser.add_argument("workbook")
    parser.add_argument("-o", "--output", default="./execution_analysis.json")
    args = parser.parse_args()
    data = parse_workbook(args.workbook)
    Path(args.output).write_text(
        json.dumps(data, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    print(
        f"✓ 执行分析完成：总计 {data['summary']['total']}，"
        f"PASS {data['summary']['pass']}，FAIL {data['summary']['fail']}，"
        f"缺陷草稿 {len(data['defects'])} → {args.output}"
    )


if __name__ == "__main__":
    main()
